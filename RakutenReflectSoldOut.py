import time
import json
import datetime
import re
import os
import csv
import base64
import requests
from openpyxl import load_workbook
import shutil
import traceback
from skpy import Skype
import pandas as pd
import xml.etree.ElementTree as ET


#PATH
MASTER = "//acad2/Ace/020_制限共用部/520_システム/WinActor/物販本社/楽天欠品作業/マスタ.xlsx"
BACKUP = os.path.abspath("backup")

def open_vault():
    f = open("config/vault.json")
    data = json.load(f)
    f.close()
    return data
    
def rakuten_authenticate_header():
    account_info = open_vault()
    data = account_info["rakuten"]["serviceSecret"] + ":" + account_info["rakuten"]["licenseKey"]
    byte_data = data.encode('utf-8')
    return {"Authorization" : "ESA " + base64.b64encode(byte_data).decode("utf-8")}

def update_stock(bulkdatas):
    auth_headers = rakuten_authenticate_header()
    auth_headers.update({"Content-Type": "application/json"})
    listed_bulkdatas = [bulkdatas[i:i + 400] for i in range(0, len(bulkdatas), 400)]

    for bulkdata in listed_bulkdatas:
        json_data = {"inventories": bulkdata}
        r = requests.post("https://api.rms.rakuten.co.jp/es/2.0/inventories/bulk-upsert", json=json_data, headers=auth_headers)
        if r.status_code != 204:
            return False
        time.sleep(1) 
    return True
        
def backup_data(path, backupdata):
    with open(path + '/upload_body.csv', 'w', newline='', encoding='utf-8-sig') as f:
        spamwriter = csv.DictWriter(f, fieldnames=["manageNumber", "variantId", "mode", "quantity"])
        spamwriter.writeheader()
        spamwriter.writerows(backupdata)

def skype_send(live_id, message):
    credentials = open_vault()
    sk = Skype(credentials["skype"]["id"], credentials["skype"]["password"])

    # 送信先の設定
    ch = sk.contacts[live_id].chat
    ch.sendMsg(message, rich=True)

def masterfile_to_dict():

    wb = load_workbook(MASTER, read_only=True)
    ws = wb["商品マスタ"]

    result = {}
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=20):
        id = row[0].value
        flag = row[19].value
        changable = row[17].value
        if not(id in result) and flag and changable != "×":
            result[id] = flag
    return result

def get_items_from_rakuten():
    auth_headers = rakuten_authenticate_header()
    auth_headers.update({"Content-Type": "application/json"})

    result = dict()
    cursorMark = ""
    nextCursorMark = "*"
    while(cursorMark != nextCursorMark):
        cursorMark = nextCursorMark
        r = requests.get("https://api.rms.rakuten.co.jp/es/2.0/items/search?hits=100&cursorMark=" + cursorMark, headers=auth_headers) #+ "&manageNumber=larmesiliconclear-30"
        if r.status_code == 200:
            for item in r.json()["results"]:
                result[item["item"]["manageNumber"]] = item["item"]
            nextCursorMark = r.json()["nextCursorMark"]

        else:
            time.sleep(1)
            cursorMark = ""
            continue

    return result

def get_stocks_from_suruzo(items):
    result = list()
    credentials = open_vault()["suruzo"]
    # to test
    credentials = credentials["test"]

    for manageNumber, details in items.items():
        r = requests.get("https://sv101.suruzo.biz/ace-1648-test/api/fukubukuro/get_stocks_control.php?login_id=" + credentials["id"] + "&password=" + credentials["password"] + "&company_product_codes=" + manageNumber)
        if r.status_code == 200:
            root = ET.fromstring(r.text.replace('&', ''))
            if root.find("STATUS").find("ERROR").text == "0":
                for item in root.find("ITEMS"):
                    if type(item.find("COLOR").text) == str:
                        color = re.search("(?<=\().+(?=\))", item.find("COLOR").text)
                        if color:
                            color_code = color.group(0)
                            for stocks in item.find("STOCKS"):
                                power = stocks.find("SIZE").text
                                if power == "0.00":
                                    power = "±0.00(度なし)"
                                site_stocks = float(stocks.find("TOKYO_STOCK").text) + float(stocks.find("FUKUOKA_STOCK").text)

                                if color_code in details and power in details[color_code]:
                                    if 0 <= site_stocks <= details[color_code]["minimum"]:
                                        set_stocks = int(site_stocks)
                                    elif site_stocks < 0:
                                        set_stocks = 0
                                    else:
                                        set_stocks = 9999
                                    result.append({
                                        "manageNumber": manageNumber,
                                        "variantId": details[color_code][power],
                                        "mode": "ABSOLUTE",
                                        "quantity": set_stocks
                                        })
    return result

def datas_for_items_request(pd_obj,verified_with_master):
    result = list()
    for index, row in pd_obj.iterrows():
        #空の場合の対応
        if pd.isna(row["自社品番"]) or pd.isna(row["カラー"]):
            continue

        managenumber = re.sub("'", "", row["自社品番"])
        if "(" in row["カラー"]:
            color = re.findall(r"(?<=\().+(?=\))", re.sub("'", "", row["カラー"]))[0]
            if not(managenumber in result) and color in verified_with_master:
                result.append(managenumber)
    return result

def filtering_with_master(items):
    result = dict()
    master_reference = masterfile_to_dict()
    for manageNumber,detail in items.items():
        if manageNumber not in result:
            result[manageNumber] = dict()
        for sku,sku_detail in detail["variants"].items():
            if "selectorValues" in sku_detail and "Key0" in sku_detail["selectorValues"] and "Key1" in sku_detail["selectorValues"]:
                color = re.search("(?<=\().+(?=\))",sku_detail["selectorValues"]["Key0"])
                if color:
                    color_code = color.group(0)
                    power = sku_detail["selectorValues"]["Key1"]
                    if color_code in master_reference:
                        if color_code not in result[manageNumber]:
                            result[manageNumber][color_code] = dict()
                            result[manageNumber][color_code]["minimum"] = master_reference[color_code]
                        result[manageNumber][color_code][power] = sku
    return result


def main(backupfolder):

    rakuten_items = get_items_from_rakuten()
    filtered_rakuten_items = filtering_with_master(rakuten_items)
    stocks_data = get_stocks_from_suruzo(filtered_rakuten_items)
    with open("test.txt", mode='w') as f:
        for row in stocks_data:
            f.write(str(row))
    #print(stocks_data)

"""     
        update_rakuten_stock(stock_info, master_reference)
        backup_data(stock_info)

"""
if __name__ == "__main__":    
    backupfolder = os.path.join(BACKUP, datetime.datetime.now().strftime("%Y%m%d%H%M%S"))
    #os.makedirs(backupfolder)

    main(backupfolder)
"""
    try:
        main(backupfolder)
    except:
        credentials = get_credentials()
        skype_send(credentials["oota"]["skypeLiveId"], "楽天欠品作業中にエラーが起きました。")
        skype_send("live:.cid.6828f41a3e80309a", "楽天欠品作業中にエラーが起きました。")
        with open(os.path.join(backupfolder, "error.log"), 'w', encoding='utf-8') as f:
            traceback.print_exc(file=f)
"""            